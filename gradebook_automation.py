import openpyxl
import math

print("hello izzy! We are working on your exam stats...")
gradebook = openpyxl.load_workbook("sample_grades.xlsx")

#define sheet
grades = gradebook.get_sheet_by_name("Sheet1")

#define test score arrays
test_1_scores = []
for i in range(2, grades.max_row + 1):
  test_1_scores.append(grades["B" + str(i)].value)

test_2_scores = []
for i in range(2, grades.max_row + 1):
  test_2_scores.append(grades["C" + str(i)].value)

test_3_scores = []
for i in range(2, grades.max_row + 1):
  test_3_scores.append(grades["D" + str(i)].value)

#define output measures
def mean(array):
  sum = 0
  for number in array:
    sum +=number
  return sum/len(array)

def median(array):
  #need to sort array first!
  array = sorted(array)
  length = len(array)
  if len(array) % 2 == 0:
    middle_location = math.ceil(length/2)
    middle = (array[middle_location] + array[middle_location-1])/2
  else:
    middle_location = length/2
    middle = array[middle_location]
  return middle

print()
print("Okay, done! The median for Exam 1 was " + str(median(test_1_scores)) + ", and the mean was " + str(mean(test_1_scores)) + ".")


